import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json, random, time, uuid, base64, io
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from supabase import create_client

# ═══════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════
SUPABASE_URL  = st.secrets["SUPABASE_URL"]
SUPABASE_KEY  = st.secrets["SUPABASE_KEY"]

@st.cache_resource
def get_sb():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

def sb():
    return get_sb()

st.set_page_config(
    page_title="Test Tizimi",
    page_icon="📝",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ═══════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════
st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding-top:1rem;max-width:720px;}
.stButton>button{width:100%;border-radius:10px;font-weight:700;height:44px;}
.stButton>button[kind="primary"]{background:linear-gradient(135deg,#1a3a6b,#2563b0);color:white;border:none;}
.card{background:white;border-radius:14px;padding:24px;box-shadow:0 4px 20px rgba(0,0,0,.1);margin-bottom:16px;}
.gerb{text-align:center;margin-bottom:8px;}
.big-title{text-align:center;font-size:26px;font-weight:800;color:#1a3a6b;margin:0;}
.sub{text-align:center;color:#888;font-size:14px;margin-bottom:20px;}
.info-bar{background:#f0f4ff;border-left:4px solid #1a3a6b;border-radius:8px;padding:10px 14px;font-size:13px;margin-bottom:12px;}
.step-badge{display:inline-block;background:#1a3a6b;color:white;font-size:12px;font-weight:700;padding:3px 12px;border-radius:20px;margin-bottom:8px;}
.result-box{text-align:center;padding:30px;background:white;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,.1);}
.pct-num{font-size:64px;font-weight:900;margin:10px 0;}
.stat-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin:16px 0;}
.stat-item{background:#f8f9fa;border-radius:10px;padding:14px;text-align:center;}
.stat-item .n{font-size:28px;font-weight:800;}
.stat-item .l{font-size:12px;color:#888;margin-top:3px;}
.q-box{background:white;border-radius:14px;padding:20px 24px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:14px;}
.q-num{font-size:12px;color:#aaa;font-weight:600;letter-spacing:.5px;margin-bottom:10px;}
.q-text{font-size:17px;font-weight:600;color:#1a3a6b;line-height:1.55;margin-bottom:20px;font-family:monospace;white-space:pre-wrap;}
.timer-ok{background:#e8f0fe;color:#1a3a6b;padding:8px 18px;border-radius:20px;font-size:18px;font-weight:800;text-align:center;}
.timer-warn{background:#fee2e2;color:#c0392b;padding:8px 18px;border-radius:20px;font-size:18px;font-weight:800;text-align:center;animation:pulse 1s infinite;}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.6}}
.guard{position:fixed;inset:0;background:#05101f;z-index:99999;display:flex;flex-direction:column;align-items:center;justify-content:center;color:white;text-align:center;padding:40px;}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════
# SCREENSHOT HIMOYA (JavaScript)
# ═══════════════════════════════════════════════════════
PROTECT_JS = """
<div id="guard" style="display:none;position:fixed;inset:0;background:#05101f;z-index:99999;
  flex-direction:column;align-items:center;justify-content:center;color:white;text-align:center;padding:40px;">
  <div style="font-size:60px;">🚫</div>
  <div style="font-size:22px;font-weight:800;margin-top:14px;">Ekran suratga olish taqiqlangan!</div>
  <div style="font-size:14px;opacity:.6;margin-top:8px;">Screenshot is not allowed during the test</div>
</div>
<script>
(function(){
  var g=document.getElementById('guard');
  function show(ms){g.style.display='flex';if(ms)setTimeout(hide,ms);}
  function hide(){g.style.display='none';}
  document.addEventListener('keyup',function(e){
    if(e.key==='PrintScreen'||e.keyCode===44){
      try{navigator.clipboard.writeText('');}catch(_){}
      show(3000);
    }
    if(e.metaKey&&e.shiftKey&&['3','4','5','6','s','S'].includes(e.key)) show(3000);
  });
  document.addEventListener('keydown',function(e){
    if(e.key==='F12'){e.preventDefault();return false;}
    if((e.ctrlKey||e.metaKey)&&e.shiftKey&&'ijcIJC'.includes(e.key)){e.preventDefault();return false;}
    if((e.ctrlKey||e.metaKey)&&(e.key==='p'||e.key==='P')){e.preventDefault();return false;}
  });
  document.addEventListener('visibilitychange',function(){
    if(document.hidden) show(0); else setTimeout(hide,800);
  });
  window.addEventListener('blur',function(){show(0);});
  window.addEventListener('focus',function(){setTimeout(hide,600);});
  document.addEventListener('contextmenu',function(e){e.preventDefault();});
  document.addEventListener('touchstart',function(e){
    if(e.touches.length>=2){e.preventDefault();show(2000);}
  },{passive:false});
  document.addEventListener('dragstart',function(e){e.preventDefault();});
})();
</script>
"""

# ═══════════════════════════════════════════════════════
# DB HELPERS
# ═══════════════════════════════════════════════════════
def db_get_teachers(include_inactive=False):
    try:
        q = sb().table('teachers').select('*').order('name')
        if not include_inactive:
            q = q.eq('is_active', 1)
        return q.execute().data or []
    except: return []

def db_get_teacher(tid):
    try:
        r = sb().table('teachers').select('*').eq('id', tid).single().execute()
        return r.data
    except: return None

def db_get_teacher_by_username(username):
    try:
        r = sb().table('teachers').select('*').eq('username', username).single().execute()
        return r.data
    except: return None

def db_save_teacher(t):
    try:
        # Bazaga yuborishdan oldin tiplarni to'g'rilash
        data = dict(t)
        data['time_limit']     = int(data.get('time_limit', 30))
        data['question_count'] = int(data.get('question_count', 10))
        data['is_active']      = int(data.get('is_active', 1))
        sb().table('teachers').upsert(data).execute()
        return True
    except Exception as e:
        st.error(f"Saqlash xatosi: {e}")
        return False

def db_delete_teacher(tid):
    try:
        sb().table('teachers').delete().eq('id', tid).execute()
        sb().table('test_files').delete().eq('teacher_id', tid).execute()
        sb().table('results').delete().eq('teacher_id', tid).execute()
        return True
    except: return False

def db_get_test_files(teacher_id):
    try:
        return sb().table('test_files').select('id,name,teacher_id').eq('teacher_id', teacher_id).execute().data or []
    except: return []

def db_save_test_file(teacher_id, name, content_bytes):
    try:
        b64 = base64.b64encode(content_bytes).decode()
        fid = str(uuid.uuid4())
        sb().table('test_files').upsert({
            'id': fid, 'teacher_id': teacher_id, 'name': name,
            'content_b64': b64, 'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }).execute()
        return True
    except: return False

def db_delete_test_file(teacher_id, name):
    try:
        sb().table('test_files').delete().eq('teacher_id', teacher_id).eq('name', name).execute()
        return True
    except: return False

def db_get_test_content(teacher_id, name):
    try:
        r = sb().table('test_files').select('content_b64').eq('teacher_id', teacher_id).eq('name', name).single().execute()
        if r.data:
            return base64.b64decode(r.data['content_b64'])
        return None
    except: return None

def db_save_result(res_data):
    try:
        sb().table('results').insert(res_data).execute()
        return True
    except: return False

def db_get_results(teacher_id=None, category=None):
    try:
        q = sb().table('results').select('*').order('date', desc=True)
        if teacher_id: q = q.eq('teacher_id', teacher_id)
        if category and category != 'all': q = q.eq('category', category)
        return q.execute().data or []
    except: return []

def db_save_active_test(token, data):
    try:
        sb().table('active_tests').upsert(data).execute()
        return True
    except: return False

def db_get_active_test(token):
    try:
        r = sb().table('active_tests').select('*').eq('token', token).single().execute()
        if r.data:
            d = r.data
            d['questions'] = json.loads(d['questions']) if isinstance(d['questions'], str) else d['questions']
            d['answers']   = json.loads(d['answers'])   if isinstance(d['answers'], str)   else d['answers']
            return d
        return None
    except: return None

def db_update_answers(token, answers):
    try:
        sb().table('active_tests').update({'answers': json.dumps(answers, ensure_ascii=False)}).eq('token', token).execute()
    except: pass

def db_delete_active_test(token):
    try:
        sb().table('active_tests').delete().eq('token', token).execute()
    except: pass

def db_get_config(key):
    try:
        r = sb().table('config').select('value').eq('key', key).single().execute()
        return r.data['value'] if r.data else None
    except: return None

def db_set_config(key, value):
    try:
        sb().table('config').upsert({'key': key, 'value': value}).execute()
        return True
    except: return False

def check_admin_pw(raw):
    h = db_get_config('admin_password')
    if not h: return raw == 'admin123'
    if h.startswith('scrypt') or h.startswith('pbkdf2'):
        try: return check_password_hash(h, raw)
        except: return False
    return h == raw

def set_admin_pw(raw):
    db_set_config('admin_password', generate_password_hash(raw))

# ═══════════════════════════════════════════════════════
# TEST HELPERS
# ═══════════════════════════════════════════════════════
def load_questions(content_bytes):
    try:
        df = pd.read_excel(io.BytesIO(content_bytes), header=None)
        questions = []
        for _, row in df.iterrows():
            vals = [str(v).strip() for v in row.tolist() if str(v).strip() not in ('nan','')]
            if len(vals) >= 2:
                questions.append({'question': vals[0], 'correct': vals[1], 'options': vals[1:]})
        return questions
    except: return []

def get_categories(teacher_id):
    files = db_get_test_files(teacher_id)
    return [f['name'] for f in files]

def get_on_categories(cats):
    on1 = [k for k in cats if '1-' in k and ('ON' in k.upper() or 'ОН' in k)]
    on2 = [k for k in cats if '2-' in k and ('ON' in k.upper() or 'ОН' in k)]
    return on1, on2

# HTML teglarni escape — matn sifatida ko'rsatish
def escape_text(s):
    return str(s) if s is not None else ''

# ═══════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════
def init_state():
    defaults = {
        'page': 'home',
        'test_token': None,
        'teacher_logged': False,
        'teacher_id': None,
        'teacher_name': None,
        'admin_logged': False,
        'q_idx': 0,
        'answered': False,
        'last_correct': None,
        'last_correct_answer': None,
        'test_finished': False,
        'test_result': None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ═══════════════════════════════════════════════════════
# PAGE: HOME
# ═══════════════════════════════════════════════════════
def page_home():
    # Admin tugmasi
    col_a, col_b = st.columns([4,1])
    with col_b:
        if st.button("⚙️ Admin", key="btn_admin_top"):
            st.session_state.page = 'admin_login'
            st.rerun()

    # Gerb + sarlavha
    st.markdown("""
    <div class="gerb">
      <svg width="80" height="80" viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg">
        <circle cx="100" cy="100" r="98" fill="#1a5276" stroke="#c9a84c" stroke-width="3"/>
        <circle cx="100" cy="100" r="82" fill="#1a6b3a" stroke="#c9a84c" stroke-width="2"/>
        <g fill="#f0d060" opacity="0.9">
          <polygon points="100,18 104,42 100,38 96,42"/>
          <polygon points="145,28 134,50 131,45 142,24"/>
          <polygon points="178,60 158,75 156,70 176,55"/>
          <polygon points="188,105 164,105 165,100 189,100"/>
          <polygon points="178,148 158,133 161,129 181,144"/>
          <polygon points="100,188 96,164 100,166 104,164"/>
          <polygon points="22,60 42,75 44,70 24,55"/>
          <polygon points="12,105 36,105 35,100 11,100"/>
          <polygon points="22,148 42,133 39,129 19,144"/>
        </g>
        <path d="M100 55 A30 30 0 1 1 100 145 A22 22 0 1 0 100 55Z" fill="#f0d060"/>
        <text x="100" y="168" text-anchor="middle" font-size="8" fill="white" font-weight="bold">O'ZBEKISTON</text>
      </svg>
    </div>
    <div class="big-title">📝 Test Tizimi</div>
    <div class="sub">O'zbekiston Respublikasi</div>
    """, unsafe_allow_html=True)

    teachers = db_get_teachers()

    # ① F.I.O. va Guruh
    st.markdown('<div class="step-badge">① Talaba ma\'lumotlari</div>', unsafe_allow_html=True)
    fio   = st.text_input("F.I.O. (To'liq ism sharif)", placeholder="Karimov Alisher Bekovich", key="inp_fio")
    group = st.text_input("Guruh", placeholder="CS-101", key="inp_group")

    step1_ok = len(fio.strip()) >= 3 and len(group.strip()) >= 1

    st.divider()

    # ② O'qituvchi
    st.markdown('<div class="step-badge">② O\'qituvchini tanlang</div>', unsafe_allow_html=True)
    teacher_names = ["— O'qituvchini tanlang —"] + [t['name'] for t in teachers]
    t_sel = st.selectbox("O'qituvchi", teacher_names, disabled=not step1_ok, key="sel_teacher")

    selected_teacher = None
    cats = []
    if t_sel != "— O'qituvchini tanlang —":
        selected_teacher = next((t for t in teachers if t['name'] == t_sel), None)
        if selected_teacher:
            st.markdown(f"""<div class="info-bar">
                ⏱ Vaqt: <b>{selected_teacher.get('time_limit',30)} daqiqa</b> &nbsp;|&nbsp;
                📋 Savollar: <b>{selected_teacher.get('question_count',10)} ta</b>
            </div>""", unsafe_allow_html=True)
            cats = get_categories(selected_teacher['id'])
            on1, on2 = get_on_categories(cats)
            if on1 and on2:
                cats.append('YaN')

    step2_ok = selected_teacher is not None

    st.divider()

    # ③ Test turi
    st.markdown('<div class="step-badge">③ Test turini tanlang</div>', unsafe_allow_html=True)
    cat_options = ["— Test turini tanlang —"] + cats
    cat_sel = st.selectbox("Test turi", cat_options, disabled=not step2_ok, key="sel_cat")
    step3_ok = cat_sel != "— Test turini tanlang —"

    st.divider()

    # Boshlash tugmasi
    can_start = step1_ok and step2_ok and step3_ok
    if st.button("🚀 Testni Boshlash", type="primary", disabled=not can_start):
        with st.spinner("Yuklanmoqda..."):
            _start_test(fio.strip(), group.strip(), selected_teacher, cat_sel)

    # O'qituvchi havolalari
    c1, c2 = st.columns(2)
    with c1:
        if st.button("👨‍🏫 O'qituvchi kirish"):
            st.session_state.page = 'teacher_login'
            st.rerun()
    with c2:
        if st.button("✏️ Ro'yxatdan o'tish"):
            st.session_state.page = 'teacher_register'
            st.rerun()

def _start_test(fio, group, teacher, category):
    tid = teacher['id']
    cats_raw = db_get_test_files(tid)
    cats_map = {f['name']: f for f in cats_raw}
    # DB dan yangi sozlamalarni olish
    fresh_teacher = db_get_teacher(tid) or teacher
    t_limit = int(fresh_teacher.get('time_limit', 30))
    q_count = int(fresh_teacher.get('question_count', 10))
    questions = []

    if category == 'YaN':
        cat_names = [f['name'] for f in cats_raw]
        on1_list, on2_list = get_on_categories(cat_names)
        if not on1_list or not on2_list:
            st.error("YaN uchun 1-ON va 2-ON fayllari zarur!"); return
        half = q_count // 2
        b1 = db_get_test_content(tid, on1_list[0])
        b2 = db_get_test_content(tid, on2_list[0])
        q1 = load_questions(b1); q2 = load_questions(b2)
        random.shuffle(q1); random.shuffle(q2)
        questions = q1[:half] + q2[q_count-half:]
    else:
        if category not in cats_map:
            st.error("Kategoriya topilmadi!"); return
        content = db_get_test_content(tid, category)
        if not content:
            st.error("Test fayli topilmadi!"); return
        all_q = load_questions(content)
        random.shuffle(all_q)
        questions = all_q[:q_count]

    for q in questions:
        opts = q['options'][:]
        random.shuffle(opts)
        q['shuffled_options'] = opts

    token = str(uuid.uuid4())
    test_id = str(uuid.uuid4())
    now = time.time()
    data = {
        'token': token, 'fio': fio, 'grp': group, 'category': category,
        'questions': json.dumps(questions, ensure_ascii=False),
        'start_time': now, 'time_limit': t_limit * 60,
        'answers': '[]', 'created_at': now, 'teacher_id': tid
    }
    db_save_active_test(token, data)
    st.session_state.test_token      = token
    st.session_state.q_idx           = 0
    st.session_state.answered        = False
    st.session_state.last_correct    = None
    st.session_state.last_correct_answer = None
    st.session_state.test_finished   = False
    st.session_state.test_result     = None
    st.session_state.page            = 'test'
    st.rerun()

# ═══════════════════════════════════════════════════════
# PAGE: TEST
# ═══════════════════════════════════════════════════════
def page_test():
    st.components.v1.html(PROTECT_JS, height=0)

    # Natija ko'rsatish (test tugagan)
    if st.session_state.test_finished:
        _show_result()
        return

    token = st.session_state.test_token
    if not token:
        st.session_state.page = 'home'; st.rerun(); return

    t = db_get_active_test(token)
    if not t:
        # Token yo'q — lekin natija session da saqlanganmi?
        if st.session_state.test_result:
            st.session_state.test_finished = True
            _show_result()
        else:
            st.session_state.page = 'home'
            st.rerun()
        return

    # Vaqtni hisoblash
    elapsed   = time.time() - float(t['start_time'])
    remaining = int(t['time_limit'] - elapsed)

    if remaining <= 0:
        _finish_test(token, t)
        return

    questions = t['questions']
    answers   = t['answers'] if isinstance(t['answers'], list) else []
    idx       = st.session_state.q_idx
    total     = len(questions)

    if idx >= total:
        _finish_test(token, t)
        return

    q = questions[idx]

    # ── HEADER ──────────────────────────────────────────
    m = remaining // 60
    s = remaining % 60
    timer_color = "#c0392b" if remaining < 60 else "#1a3a6b"
    prog = int(idx / total * 100)

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1a3a6b,#2563b0);color:white;
      padding:12px 18px;border-radius:12px;margin-bottom:10px;
      display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">
      <div>
        <div style="font-weight:700;font-size:15px;">👤 {t['fio']}</div>
        <div style="font-size:12px;opacity:.8;">Guruh: {t['grp']} | {t['category']}</div>
      </div>
      <div style="background:rgba(255,255,255,.2);border-radius:8px;padding:7px 16px;
        font-size:20px;font-weight:800;min-width:80px;text-align:center;
        {'animation:pulse 1s infinite;background:#c0392b;' if remaining < 60 else ''}">
        {m:02d}:{s:02d}
      </div>
    </div>
    <div style="background:#dde;height:6px;border-radius:3px;margin-bottom:14px;">
      <div style="background:linear-gradient(90deg,#1a3a6b,#2563b0);
        height:6px;border-radius:3px;width:{prog}%;"></div>
    </div>
    """, unsafe_allow_html=True)

    # ── SAVOL ───────────────────────────────────────────
    st.markdown(f"""
    <div class="q-box">
      <div class="q-num">SAVOL {idx+1} / {total}</div>
      <div class="q-text">{escape_text(q['question'])}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── JAVOBLAR ────────────────────────────────────────
    LETTERS = ['A','B','C','D','E']

    if not st.session_state.answered:
        # Javob tanlanmagan — variantlarni tugma sifatida ko'rsatish
        for j, opt in enumerate(q['shuffled_options']):
            if st.button(f"{LETTERS[j]}) {escape_text(opt)}", key=f"opt_{idx}_{j}"):
                correct = (opt == q['correct'])
                answers.append({'index': idx, 'answer': opt, 'correct': correct})
                db_update_answers(token, answers)
                st.session_state.answered             = True
                st.session_state.last_correct         = correct
                st.session_state.last_correct_answer  = q['correct']
                st.rerun()

        # Vaqt yangilash
        st.caption(f"⏱ Qolgan vaqt: {m:02d}:{s:02d}")
        if st.button("🔄 Vaqtni yangilash", key=f"ref_{idx}"):
            st.rerun()

    else:
        # Javob tanlangan — to'g'ri/noto'g'rini ko'rsatish
        chosen = answers[-1]['answer'] if answers else None
        for j, opt in enumerate(q['shuffled_options']):
            is_correct = (opt == q['correct'])
            is_chosen  = (opt == chosen)
            if is_correct and is_chosen:
                st.success(f"✅ {LETTERS[j]}) {escape_text(opt)}")
            elif is_correct:
                st.success(f"✅ {LETTERS[j]}) {escape_text(opt)}")
            elif is_chosen:
                st.error(f"❌ {LETTERS[j]}) {escape_text(opt)}")
            else:
                st.markdown(f"&nbsp;&nbsp;&nbsp;{LETTERS[j]}) {escape_text(opt)}")

        # Noto'g'ri bo'lsa faqat to'g'ri javobni ko'rsat
        if not st.session_state.last_correct:
            st.caption(f"💡 To\'g\'ri javob: {escape_text(st.session_state.last_correct_answer)}")


        st.write("")
        btn_label = "Yakunlash ✓" if idx >= total - 1 else "Keyingisi ›"
        if st.button(btn_label, type="primary", key=f"next_{idx}"):
            if idx >= total - 1:
                t2 = db_get_active_test(token)
                _finish_test(token, t2 or t)
            else:
                st.session_state.q_idx    += 1
                st.session_state.answered  = False
                st.session_state.last_correct = None
                st.session_state.last_correct_answer = None
                st.rerun()

def _finish_test(token, t):
    answers = t['answers'] if isinstance(t['answers'], list) else json.loads(t['answers'])
    total   = len(t['questions'])
    correct = sum(1 for a in answers if a['correct'])
    pct     = round(correct / total * 100, 1) if total > 0 else 0
    result  = {
        'id': str(uuid.uuid4()), 'fio': t['fio'], 'grp': t['grp'],
        'category': t['category'], 'total': total, 'correct': correct,
        'wrong': total - correct, 'percentage': pct,
        'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'answers': json.dumps(answers, ensure_ascii=False),
        'teacher_id': t.get('teacher_id')
    }
    db_save_result(result)
    db_delete_active_test(token)
    st.session_state.test_finished = True
    st.session_state.test_result   = result
    st.rerun()

def _show_result():
    r = st.session_state.test_result
    if not r: return
    pct = r['percentage']
    if pct >= 85:   icon,grade,color = "🏆","A'lo (Excellent)","#27ae60"
    elif pct >= 70: icon,grade,color = "🌟","Yaxshi (Good)","#2563b0"
    elif pct >= 55: icon,grade,color = "👍","Qoniqarli","#f39c12"
    else:           icon,grade,color = "📚","Qoniqarsiz","#e74c3c"

    st.markdown(f"""
    <div class="result-box">
      <div style="font-size:60px;">{icon}</div>
      <div style="font-size:18px;color:#aaa;margin-top:10px;">Natija</div>
      <div class="pct-num" style="color:{color};">{pct}%</div>
      <div style="font-size:18px;font-weight:700;color:{color};margin-bottom:16px;">{grade}</div>
      <div class="stat-grid">
        <div class="stat-item"><div class="n" style="color:#2563b0;">{r['total']}</div><div class="l">Jami savollar</div></div>
        <div class="stat-item"><div class="n" style="color:#27ae60;">{r['correct']}</div><div class="l">To'g'ri javob</div></div>
        <div class="stat-item"><div class="n" style="color:#e74c3c;">{r['wrong']}</div><div class="l">Xato javob</div></div>
      </div>
      <div style="color:#aaa;font-size:13px;">👤 {r['fio']} | Guruh: {r['grp']} | 📂 {r['category']}<br>📅 {r['date']}</div>
    </div>
    """, unsafe_allow_html=True)

    st.write("")
    if st.button("🏠 Bosh sahifaga qaytish", type="primary"):
        for k in ['test_token','test_finished','test_result','q_idx','answered','last_correct','last_correct_answer']:
            st.session_state[k] = None if k in ['test_token','test_result'] else False if 'answered' in k or 'finished' in k or 'correct' in k else 0
        st.session_state.page = 'home'
        st.rerun()

# ═══════════════════════════════════════════════════════
# PAGE: TEACHER LOGIN
# ═══════════════════════════════════════════════════════
def page_teacher_login():
    st.markdown('<div class="big-title">👨‍🏫 O\'qituvchi Kirish</div>', unsafe_allow_html=True)
    st.write("")
    username = st.text_input("Login", placeholder="username")
    password = st.text_input("Parol", type="password")
    if st.button("🔐 Kirish", type="primary"):
        t = db_get_teacher_by_username(username.strip())
        if t and check_password_hash(t['password_hash'], password):
            st.session_state.teacher_logged = True
            st.session_state.teacher_id     = t['id']
            st.session_state.teacher_name   = t['name']
            st.session_state.page = 'teacher_dashboard'
            st.rerun()
        else:
            st.error("❌ Login yoki parol noto'g'ri!")
    if st.button("← Bosh sahifa"):
        st.session_state.page = 'home'; st.rerun()
    st.divider()
    if st.button("✏️ Ro'yxatdan o'tish"):
        st.session_state.page = 'teacher_register'; st.rerun()

# ═══════════════════════════════════════════════════════
# PAGE: TEACHER REGISTER
# ═══════════════════════════════════════════════════════
def page_teacher_register():
    st.markdown('<div class="big-title">✏️ Ro\'yxatdan O\'tish</div>', unsafe_allow_html=True)
    st.write("")
    name     = st.text_input("To'liq ism (F.I.O.) *", placeholder="Karimov Alisher")
    c1, c2   = st.columns(2)
    with c1: username = st.text_input("Login *", placeholder="alisher_k")
    with c2: email    = st.text_input("Email", placeholder="email@mail.com")
    c3, c4   = st.columns(2)
    with c3: pw  = st.text_input("Parol *", type="password")
    with c4: pw2 = st.text_input("Parolni takrorlang *", type="password")

    if st.button("✅ Ro'yxatdan o'tish", type="primary"):
        if not all([name.strip(), username.strip(), pw]):
            st.error("❌ Barcha majburiy maydonlarni to'ldiring!")
        elif pw != pw2:
            st.error("❌ Parollar mos kelmadi!")
        elif len(pw) < 6:
            st.error("❌ Parol kamida 6 ta belgi!")
        elif db_get_teacher_by_username(username.strip()):
            st.error("❌ Bu login band!")
        else:
            tid = str(uuid.uuid4())
            t = {
                'id': tid, 'name': name.strip(), 'username': username.strip(),
                'password_hash': generate_password_hash(pw),
                'email': email.strip(), 'is_active': 1,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'time_limit': 30, 'question_count': 10
            }
            if db_save_teacher(t):
                st.session_state.teacher_logged = True
                st.session_state.teacher_id     = tid
                st.session_state.teacher_name   = name.strip()
                st.session_state.page = 'teacher_dashboard'
                st.rerun()
            else:
                st.error("❌ Xatolik yuz berdi!")
    if st.button("← Kirish sahifasi"):
        st.session_state.page = 'teacher_login'; st.rerun()

# ═══════════════════════════════════════════════════════
# PAGE: TEACHER DASHBOARD
# ═══════════════════════════════════════════════════════
def page_teacher_dashboard():
    if not st.session_state.teacher_logged:
        st.session_state.page = 'teacher_login'; st.rerun()

    teacher = db_get_teacher(st.session_state.teacher_id)
    if not teacher:
        st.session_state.page = 'teacher_login'; st.rerun()

    c1, c2 = st.columns([3,1])
    with c1:
        st.markdown(f"### 👨‍🏫 {teacher['name']}")
        st.caption(f"@{teacher['username']}")
    with c2:
        if st.button("🚪 Chiqish"):
            st.session_state.teacher_logged = False
            st.session_state.teacher_id = None
            st.session_state.page = 'home'; st.rerun()

    tab1, tab2, tab3 = st.tabs(["⚙️ Sozlamalar", "📁 Test Fayllar", "📊 Natijalar"])

    # ── SOZLAMALAR ──
    with tab1:
        st.subheader("Test Sozlamalari")
        c1, c2 = st.columns(2)
        with c1: tl = st.number_input("⏱ Vaqt limiti (daqiqa)", min_value=1, max_value=300, value=int(teacher.get('time_limit',30)))
        with c2: qc = st.number_input("📋 Savollar soni", min_value=1, max_value=200, value=int(teacher.get('question_count',10)))
        if st.button("💾 Saqlash", type="primary"):
            teacher['time_limit']     = int(tl)
            teacher['question_count'] = int(qc)
            if db_save_teacher(teacher):
                st.success(f"✅ Saqlandi! Vaqt: {tl} daqiqa, Savollar: {qc} ta")
                st.rerun()
            else:
                st.error("❌ Saqlashda xatolik!")

    # ── TEST FAYLLAR ──
    with tab2:
        st.subheader("Test Fayl Yuklash")
        st.caption("XLSX format: 1-ustun savol | 2-ustun to'g'ri javob | 3,4,5-ustunlar noto'g'ri javoblar")
        name_inp = st.text_input("Test nomi", placeholder="Masalan: 1-ON_Matematika")
        file_up  = st.file_uploader("XLSX fayl", type=['xlsx'])
        if st.button("📤 Yuklash", type="primary"):
            if not name_inp.strip() or not file_up:
                st.error("❌ Test nomi va faylni kiriting!")
            else:
                content = file_up.read()
                if db_save_test_file(teacher['id'], name_inp.strip(), content):
                    st.success(f"✅ '{name_inp}' muvaffaqiyatli yuklandi!")
                    st.rerun()
                else:
                    st.error("❌ Yuklashda xatolik!")

        st.divider()
        st.subheader("Mening Test Fayllarim")
        files = db_get_test_files(teacher['id'])
        if files:
            for f in files:
                c1, c2 = st.columns([4,1])
                with c1: st.markdown(f"📄 **{f['name']}**")
                with c2:
                    if st.button("🗑", key=f"del_{f['id']}"):
                        db_delete_test_file(teacher['id'], f['name'])
                        st.rerun()
        else:
            st.info("Hozircha test fayllaringiz yo'q")

    # ── NATIJALAR ──
    with tab3:
        st.subheader("Talabalar Natijalari")
        results = db_get_results(teacher_id=teacher['id'])
        if results:
            done   = len(results)
            passed = sum(1 for r in results if r['percentage'] >= 55)
            c1,c2,c3 = st.columns(3)
            c1.metric("Jami", done)
            c2.metric("O'tganlar (≥55%)", passed)
            c3.metric("O'tmaganlar", done-passed)

            cats = ['all'] + list(set(r['category'] for r in results))
            sel_cat = st.selectbox("Kategoriya filtri", cats)
            filtered = results if sel_cat == 'all' else [r for r in results if r['category'] == sel_cat]

            df = pd.DataFrame([{
                'F.I.O.': r['fio'], 'Guruh': r['grp'], 'Kategoriya': r['category'],
                'Jami': r['total'], "To'g'ri": r['correct'], 'Xato': r['wrong'],
                'Foiz (%)': r['percentage'], 'Sana': r['date']
            } for r in filtered])
            st.dataframe(df, use_container_width=True)

            excel_buf = _build_excel(filtered)
            st.download_button("📥 Excel Yuklab olish", excel_buf,
                file_name=f"natijalar_{sel_cat}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Hozircha natijalar yo'q")

# ═══════════════════════════════════════════════════════
# PAGE: ADMIN LOGIN
# ═══════════════════════════════════════════════════════
def page_admin_login():
    st.markdown('<div class="big-title">🛡️ Administrator Paneli</div>', unsafe_allow_html=True)
    st.write("")
    pw = st.text_input("Administrator paroli", type="password")
    if st.button("🔐 Kirish", type="primary"):
        if check_admin_pw(pw):
            st.session_state.admin_logged = True
            st.session_state.page = 'admin'
            st.rerun()
        else:
            st.error("❌ Parol noto'g'ri!")
    if st.button("← Bosh sahifa"):
        st.session_state.page = 'home'; st.rerun()

# ═══════════════════════════════════════════════════════
# PAGE: ADMIN DASHBOARD
# ═══════════════════════════════════════════════════════
def page_admin():
    if not st.session_state.admin_logged:
        st.session_state.page = 'admin_login'; st.rerun()

    c1, c2 = st.columns([3,1])
    with c1: st.markdown("### 🛡️ Administrator Paneli")
    with c2:
        if st.button("🚪 Chiqish"):
            st.session_state.admin_logged = False
            st.session_state.page = 'home'; st.rerun()

    teachers = db_get_teachers(include_inactive=True)
    results_all = db_get_results()
    done_all   = len(results_all)
    passed_all = sum(1 for r in results_all if r['percentage'] >= 55)

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Foydalanuvchilar", len(teachers))
    c2.metric("Faollar", sum(1 for t in teachers if t['is_active']))
    c3.metric("Jami topshirishlar", done_all)
    c4.metric("O'tganlar (≥55%)", passed_all)

    tab1, tab2, tab3 = st.tabs(["👥 Foydalanuvchilar", "📊 Hisobot", "🔑 Parol"])

    # ── FOYDALANUVCHILAR ──
    with tab1:
        st.subheader("Yangi Foydalanuvchi Qo'shish")
        c1,c2 = st.columns(2)
        with c1: a_name = st.text_input("To'liq ism *", key="a_name")
        with c2: a_user = st.text_input("Login *", key="a_user")
        c3,c4 = st.columns(2)
        with c3: a_email = st.text_input("Email", key="a_email")
        with c4: a_pw = st.text_input("Parol *", type="password", key="a_pw")
        c5,c6 = st.columns(2)
        with c5: a_tl = st.number_input("Vaqt (daqiqa)", value=30, min_value=1, max_value=300, key="a_tl")
        with c6: a_qc = st.number_input("Savollar soni", value=10, min_value=1, max_value=200, key="a_qc")

        if st.button("➕ Qo'shish", type="primary"):
            if not all([a_name.strip(), a_user.strip(), a_pw]):
                st.error("❌ Majburiy maydonlarni to'ldiring!")
            elif len(a_pw) < 6:
                st.error("❌ Parol kamida 6 ta belgi!")
            elif db_get_teacher_by_username(a_user.strip()):
                st.error("❌ Bu login band!")
            else:
                tid = str(uuid.uuid4())
                t = {'id':tid,'name':a_name.strip(),'username':a_user.strip(),
                     'password_hash':generate_password_hash(a_pw),
                     'email':a_email.strip(),'is_active':1,
                     'created_at':datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                     'time_limit':a_tl,'question_count':a_qc}
                if db_save_teacher(t): st.success("✅ Qo'shildi!"); st.rerun()
                else: st.error("❌ Xatolik!")

        st.divider()
        st.subheader("Foydalanuvchilar Ro'yxati")
        if teachers:
            for t in teachers:
                t_results = [r for r in results_all if r.get('teacher_id') == t['id']]
                t_files   = db_get_test_files(t['id'])
                status    = "🟢 Faol" if t['is_active'] else "🔴 Bloklangan"
                with st.expander(f"{status} | {t['name']} (@{t['username']}) — {len(t_files)} fayl, {len(t_results)} natija"):
                    c1,c2,c3 = st.columns(3)
                    with c1: st.write(f"📧 {t.get('email') or '—'}")
                    with c2: st.write(f"⏱ {t.get('time_limit',30)} daqiqa / {t.get('question_count',10)} ta savol")
                    with c3: st.write(f"📅 {(t.get('created_at') or '')[:10]}")

                    # Tahrirlash
                    with st.form(key=f"edit_{t['id']}"):
                        ec1,ec2 = st.columns(2)
                        with ec1: e_name = st.text_input("Ism", value=t['name'], key=f"en_{t['id']}")
                        with ec2: e_user = st.text_input("Login", value=t['username'], key=f"eu_{t['id']}")
                        ec3,ec4 = st.columns(2)
                        with ec3: e_tl = st.number_input("Vaqt", value=int(t.get('time_limit',30)), key=f"etl_{t['id']}")
                        with ec4: e_qc = st.number_input("Savollar", value=int(t.get('question_count',10)), key=f"eqc_{t['id']}")
                        e_pw = st.text_input("Yangi parol (bo'sh = o'zgarmaydi)", type="password", key=f"epw_{t['id']}")
                        if st.form_submit_button("💾 Saqlash"):
                            t['name'] = e_name; t['username'] = e_user
                            t['time_limit'] = e_tl; t['question_count'] = e_qc
                            if e_pw and len(e_pw) >= 6:
                                t['password_hash'] = generate_password_hash(e_pw)
                            if db_save_teacher(t): st.success("✅ Saqlandi!"); st.rerun()
                            else: st.error("❌ Xatolik!")

                    c_btn1, c_btn2, c_btn3 = st.columns(3)
                    with c_btn1:
                        lbl = "🔒 Bloklash" if t['is_active'] else "🔓 Ochish"
                        if st.button(lbl, key=f"tog_{t['id']}"):
                            t['is_active'] = 0 if t['is_active'] else 1
                            db_save_teacher(t); st.rerun()
                    with c_btn2:
                        if t_results:
                            buf = _build_excel(t_results)
                            st.download_button("📥 Natijalar", buf,
                                file_name=f"natijalar_{t['username']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"exp_{t['id']}")
                    with c_btn3:
                        if st.button("🗑 O'chirish", key=f"del_{t['id']}"):
                            db_delete_teacher(t['id']); st.rerun()
        else:
            st.info("Hozircha foydalanuvchilar yo'q")

    # ── HISOBOT ──
    with tab2:
        st.subheader("Umumiy Hisobot")
        rows = []
        for t in teachers:
            t_results = [r for r in results_all if r.get('teacher_id') == t['id']]
            t_files   = db_get_test_files(t['id'])
            done   = len(t_results)
            passed = sum(1 for r in t_results if r['percentage'] >= 55)
            rows.append({
                'Foydalanuvchi': t['name'],
                'Login': t['username'],
                'Fan / Test fayllari': ', '.join(f['name'] for f in t_files) or '—',
                'Testlar soni': len(t_files),
                'Ishlanganlar': done,
                "O'tganlar (≥55%)": passed,
                "O'tmaganlar": done - passed,
                "O'tish (%)": round(passed/done*100,1) if done > 0 else 0
            })
        if rows:
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True)
            buf = _build_report_excel(rows)
            st.download_button("📥 Excel Yuklab olish", buf,
                file_name=f"hisobot_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── PAROL ──
    with tab3:
        st.subheader("Administrator Parolini O'zgartirish")
        cur = st.text_input("Joriy parol", type="password", key="adm_cur")
        nw  = st.text_input("Yangi parol (kamida 6 belgi)", type="password", key="adm_new")
        nw2 = st.text_input("Yangi parolni takrorlang", type="password", key="adm_new2")
        if st.button("💾 Yangilash", type="primary"):
            if not check_admin_pw(cur): st.error("❌ Joriy parol noto'g'ri!")
            elif len(nw) < 6: st.error("❌ Yangi parol kamida 6 ta belgi!")
            elif nw != nw2: st.error("❌ Parollar mos kelmadi!")
            else:
                set_admin_pw(nw); st.success("✅ Parol yangilandi!")

# ═══════════════════════════════════════════════════════
# EXCEL BUILDERS
# ═══════════════════════════════════════════════════════
def _build_excel(results):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Natijalar"
    hf = PatternFill("solid",fgColor="1A3A6B")
    hfont = Font(bold=True,color="FFFFFF",name="Arial",size=11)
    border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ef = PatternFill("solid",fgColor="EBF5FB")
    ws.append(["№","F.I.O.","Guruh","Kategoriya","Jami","To'g'ri","Xato","Foiz (%)","Sana"])
    for cell in ws[1]:
        cell.fill=hf;cell.font=hfont;cell.alignment=Alignment(horizontal='center',vertical='center');cell.border=border
    ws.row_dimensions[1].height=26
    for i,r in enumerate(results,1):
        ws.append([i,r.get('fio',''),r.get('grp',''),r.get('category',''),
                   r.get('total',0),r.get('correct',0),r.get('wrong',0),r.get('percentage',0),r.get('date','')])
        for cell in ws[ws.max_row]:
            cell.border=border;cell.alignment=Alignment(horizontal='center',vertical='center')
            if i%2==0:cell.fill=ef
        pct=ws.cell(row=ws.max_row,column=8)
        if isinstance(pct.value,(int,float)):
            pct.font=Font(color="1E8449" if pct.value>=85 else("D4AC0D" if pct.value>=55 else"C0392B"),bold=True,name="Arial")
    for i,w in enumerate([5,28,14,18,12,12,10,12,20],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf

def _build_report_excel(rows):
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Hisobot"
    hf=PatternFill("solid",fgColor="1A3A6B")
    hfont=Font(bold=True,color="FFFFFF",name="Arial",size=11)
    border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ef=PatternFill("solid",fgColor="EBF5FB")
    headers=list(rows[0].keys()) if rows else []
    ws.append(headers)
    for cell in ws[1]:
        cell.fill=hf;cell.font=hfont;cell.alignment=Alignment(horizontal='center',vertical='center');cell.border=border
    ws.row_dimensions[1].height=26
    for i,row in enumerate(rows,1):
        ws.append(list(row.values()))
        for cell in ws[ws.max_row]:
            cell.border=border;cell.alignment=Alignment(horizontal='center',vertical='center')
            if i%2==0:cell.fill=ef
    for i in range(1,len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=20
    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf

# ═══════════════════════════════════════════════════════
# ROUTER
# ═══════════════════════════════════════════════════════
page = st.session_state.page
if   page == 'home':               page_home()
elif page == 'test':               page_test()
elif page == 'teacher_login':      page_teacher_login()
elif page == 'teacher_register':   page_teacher_register()
elif page == 'teacher_dashboard':  page_teacher_dashboard()
elif page == 'admin_login':        page_admin_login()
elif page == 'admin':              page_admin()
else:
    st.session_state.page = 'home'; st.rerun()
